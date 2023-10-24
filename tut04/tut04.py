#Help https://youtu.be/H37f_x4wAC0

from array import array
from itertools import count
from math import radians
from operator import mod

import pandas as pd
def octant_longest_subsequence_count_with_range():
###Code




    df=pd.read_excel('input_octant_longest_subsequence_with_range.xlsx')
    
    
    avg_u=df['U'].mean()
    avg_v=df['V'].mean()
    avg_w=df['W'].mean()
    df.loc[0,"U avg"]=avg_u
    df.loc[0,"V avg"]=avg_v
    df.loc[0,"W avg"]=avg_w

    di1=[]
    di2=[]
    di3=[]

    for i in df['U']:
        di1.append(i-avg_u)
    for i in df['V']:
        di2.append(i-avg_v)
    for i in df['W']:
        di3.append(i-avg_w)

    df["U' = U - Avg"]=di1
    df["V' = V - Avg"]=di2
    df["W' = W - Avg"]=di3

    oct=[]

    cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
    for i in range(len(df['U'])):
        if(di1[i]>0 and di2[i]>0):
            if(di3[i]>0):
                oct.append(1)
                cp1+=1
            else:
                oct.append(-1)
                cn1+=1
        elif(di1[i]<0 and di2[i]>0):
            if(di3[i]>0):
                oct.append(2)
                cp2+=1
            else:
                oct.append(-2)
                cn2+=1
        
        elif(di1[i]<0 and di2[i]<0):
            if(di3[i]>0):
                oct.append(3)
                cp3+=1
            else:
                oct.append(-3)
                cn3+=1
        
        elif(di1[i]>0 and di2[i]<0):
            if(di3[i]>0):
                oct.append(4)
                cp4+=1
            else:
                oct.append(-4)

    df["octant"]=oct
    cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
    mp1=mn1=mp2=mn2=mp3=mn3=mn4=mp4=0
    c_p1=c_n1=c_p2=c_n2=c_p3=c_n3=c_p4=c_n4=0

    for i in df['octant']:
        if(i==1):
            cp1+=1
        else:
            mp1=max(cp1,mp1)
            cp1=0
        if i==-1:
            cn1+=1
        else:
            mn1=max(cn1,mn1)
            cn1=0
        if(i==2):
            cp2+=1
        else:
            mp2=max(cp2,mp2)
            cp2=0

        if(i==-2):
            cn2+=1
        else:
            mn2=max(cn2,mn2)
            cn2=0

        if(i==3):
            cp3+=1
        else:
            mp3=max(cp3,mp3)
            cp3=0

        if(i==-3):
            cn3+=1
        else:
            mn3=max(cn3,mn3)
            cn3=0

        if(i==4):
            cp4+=1
        else:
            mp4=max(cp4,mp4)
            cp4=0
        if(i==-4):
            cn4+=1
        else:
            mn4=max(cn4,mn4)
            cn4=0
    time_p1=[]
    time_n1=[]
    time_p2=[]
    time_n2=[]
    time_p3=[]
    time_n3=[]
    time_p4=[]
    time_n4=[]
    
    k=0
    for i in df['octant']:
        if i==1:
            cp1+=1
        else:
            if cp1==mp1:
                c_p1+=1
                time_p1.append([df.Time[k-mp1],df.Time[k-1]])
            cp1=0
            
        if i==-1:
            cn1+=1
        else:
            if cn1==mn1:
                c_n1+=1
                time_n1.append([df.Time[k-mn1],df.Time[k-1]])
            cn1=0
                

        

        if i==2:
            cp2+=1
        else:
            if cp2==mp2:
                c_p2+=1
                time_p2.append([df.Time[k-mp2],df.Time[k-1]])
            cp2=0

        if i==-2:

            cn2+=1
        else:
            if cn2==mn2:
                c_n2+=1
                time_n2.append([df.Time[k-mn2],df.Time[k-1]])
            cn2=0

        if i==3:

            cp3+=1
        else:
            if cp3==mp3:
                c_p3+=1
                time_p3.append([df.Time[k-mp3],df.Time[k-1]])
            cp3=0  

        if i==-3:
            cn3+=1
        else:
            if cn3==mn3:
                c_n3+=1
                time_n3.append([df.Time[k-mn3],df.Time[k-1]])
            cn3=0


        if i==4:
            cp4+=1
        else:
            if cp4==mp4:
                c_p4+=1
                time_p4.append([df.Time[k-mp4],df.Time[k-1]])
            cp4=0

        if i==-4:
            cn4+=1
        else:
            if cn4==mn4:
                c_n4+=1
                time_n4.append([df.Time[k-mn4],df.Time[k-1]])
            cn4=0
        
        k+=1


    twoD_array=[[0 for i in range(2)] for i in range(8)]
    twoD_array[0][0]=mp1
    twoD_array[0][1]=c_p1
    twoD_array[1][0]=mn1
    twoD_array[1][1]=c_n1
    twoD_array[2][0]=mp2
    twoD_array[2][1]=c_p2
    twoD_array[3][0]=mn2
    twoD_array[3][1]=c_n2
    twoD_array[4][0]=mp3
    twoD_array[4][1]=c_p3
    twoD_array[5][0]=mn3
    twoD_array[5][1]=c_n3
    twoD_array[6][0]=mp4
    twoD_array[6][1]=c_p4
    twoD_array[7][0]=mn4
    twoD_array[7][1]=c_n4

    df.loc[0,""]=""
    header=["Count","+1","-1","+2","-2","+3","-3","+4","-4"]   ## header defined 
    header1=["Longest Subsquence Length","Count"]

    ##traversing header
    for i in range(len(header)):
        df.loc[i+1," "]=header[i]
        if i==0:
            df.loc[i+1,"  "]=header1[0]
            df.loc[i+1,"   "]=header1[1]
        if i!=0:
            df.loc[i+1,"  "]=twoD_array[i-1][0]
            df.loc[i+1,"   "]=twoD_array[i-1][1]

    df.to_excel("output_octant_longest_subsequence_with_range.xlsx")



    from openpyxl import load_workbook
    wb = load_workbook(r'output_octant_longest_subsequence_with_range.xlsx')

    Sheet1 = wb.active
    Sheet1['R3']='Count'
    Sheet1['S3']='Longest Subsquence Length'
    Sheet1['T3']='Count'

    octat=[1,-1,2,-2,3,-3,4,-4]
    i=4
    for oc in octat:
        if(oc==1):
            Sheet1.cell(row=i,column=18).value='+1'
            Sheet1.cell(row=i,column=19).value=mp1
            Sheet1.cell(row=i,column=20).value=c_p1
            Sheet1.cell(row=i+1,column=18).value='     Time'
            Sheet1.cell(row=i+1,column=19).value='     To'
            Sheet1.cell(row=i+1,column=20).value='     From'

            r=0
            while(r<c_p1):
                Sheet1.cell(row=i+2+r,column=19).value=time_p1[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_p1[r][1]
                r+=1

        if(oc==-1):
            Sheet1.cell(row=i,column=18).value='-1'
            Sheet1.cell(row=i,column=19).value=mn1
            Sheet1.cell(row=i,column=20).value=c_n1
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_n1):
                Sheet1.cell(row=i+2+r,column=19).value=time_n1[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_n1[r][1]
                r+=1

        if(oc==2):
            Sheet1.cell(row=i,column=18).value='+2'
            Sheet1.cell(row=i,column=19).value=mp2
            Sheet1.cell(row=i,column=20).value=c_p2
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_p2):
                Sheet1.cell(row=i+2+r,column=19).value=time_p2[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_p2[r][1]
                r+=1

        if(oc==-2):
            Sheet1.cell(row=i,column=18).value='-2'
            Sheet1.cell(row=i,column=19).value=mn2
            Sheet1.cell(row=i,column=20).value=c_n2
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_n2):
                Sheet1.cell(row=i+2+r,column=19).value=time_n2[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_n2[r][1]
                r+=1

        if(oc==3):
            Sheet1.cell(row=i,column=18).value='+3'
            Sheet1.cell(row=i,column=19).value=mp3
            Sheet1.cell(row=i,column=20).value=c_p3
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_p3):
                Sheet1.cell(row=i+2+r,column=19).value=time_p3[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_p3[r][1]
                r+=1

        if(oc==-3):
            Sheet1.cell(row=i,column=18).value='-3'
            Sheet1.cell(row=i,column=19).value=mn3
            Sheet1.cell(row=i,column=20).value=c_n3
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_n3):
                Sheet1.cell(row=i+2+r,column=19).value=time_n3[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_n3[r][1]
                r+=1

        if(oc==4):
            Sheet1.cell(row=i,column=18).value='+4'
            Sheet1.cell(row=i,column=19).value=mp4
            Sheet1.cell(row=i,column=20).value=c_p4
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_p4):
                Sheet1.cell(row=i+2+r,column=19).value=time_p4[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_p4[r][1]
                r+=1
        if(oc==-4):
            Sheet1.cell(row=i,column=18).value='-4'
            Sheet1.cell(row=i,column=19).value=mn4
            Sheet1.cell(row=i,column=20).value=c_n4
            Sheet1.cell(row=i+1,column=18).value='      Time'
            Sheet1.cell(row=i+1,column=19).value='      To'
            Sheet1.cell(row=i+1,column=20).value='      From'

            r=0
            while(r<c_n4):
                Sheet1.cell(row=i+2+r,column=19).value=time_n4[r][0]
                Sheet1.cell(row=i+2+r,column=20).value=time_n4[r][1]
                r+=1
        i+=2+r
    

            

    wb.save('output_octant_longest_subsequence_with_range.xlsx')
        



octant_longest_subsequence_count_with_range()