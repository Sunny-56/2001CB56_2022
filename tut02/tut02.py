from array import array
from operator import mod
import pandas as pd
df=pd.read_excel('‌input_trial.xlsx')
# df.drop(df.columns[4:26],axis=1,inplace=True)
mod=5000
avg_u=df['U'].mean()
avg_v=df['V'].mean()
avg_w=df['W'].mean()
df.loc[0,"U avg"]=avg_u
df.loc[0,"V avg"]=avg_v
df.loc[0,"W avg"]=avg_w

di1=[]
di2=[]
di3=[]

for i in df['U']:
    di1.append(i-avg_u)
for i in df['V']:
    di2.append(i-avg_v)
for i in df['W']:
    di3.append(i-avg_w)

df["U' = U - Avg"]=di1
df["V' = V - Avg"]=di2
df["W' = W - Avg"]=di3

oct=[]
cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
for i in range(len(df['U'])):
    if(di1[i]>0 and di2[i]>0):
        if(di3[i]>0):
            oct.append(1)
            cp1+=1
        else:
            oct.append(-1)
            cn1+=1
    elif(di1[i]<0 and di2[i]>0):
        if(di3[i]>0):
            oct.append(2)
            cp2+=1
        else:
            oct.append(-2)
            cn2+=1
    
    elif(di1[i]<0 and di2[i]<0):
        if(di3[i]>0):
            oct.append(3)
            cp3+=1
        else:
            oct.append(-3)
            cn3+=1
    
    elif(di1[i]>0 and di2[i]<0):
        if(di3[i]>0):
            oct.append(4)
            cp4+=1
        else:
            oct.append(-4)

df["octant"]=oct
df.loc[2,'']='User input'


list_p1=[]
list_n1=[]
list_p2=[]
list_n2=[]
list_p3=[]
list_n3=[]
list_p4=[]
list_n4=[]

list_p1.append(cp1)
list_p1.append("")
list_n1.append(cn1)
list_n1.append('')
list_p2.append(cp2)
list_p2.append('')
list_n2.append(cn2)
list_n2.append('')
list_p3.append(cp3)
list_p3.append('')
list_n3.append(cn3)
list_n3.append('')
list_p4.append(cp4)
list_p4.append('')
list_n4.append(cn4)
list_n4.append('')

oct_id=['Overall count',f"mod-{mod}"]
flag=count=0
while True:
    cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
    num=count+mod
    nI=1+count
    if(count+mod)>29745:
        num=29745
    if(count==0):
        nI=0
    oct_id.append(f"{nI-1}-{num-1}")

    for i in range(count,count+mod):
        if(i>=len(df['U'])):
            flag=1
            break
        if(oct[i]==1):
            cp1+=1
        if(oct[i]==-1):
            cn1+=1
        if(oct[i]==2):
            cp2+=1
        if(oct[i]==-2):
            cn2+=1
        if(oct[i]==3):
            cp3+=1
        if(oct[i]==-3):
            cn3+=1
        if(oct[i]==4):
            cp4+=4
        if(oct[i]==-4):
            cn4+=1
        count+=1
    

    list_p1.append(cp1)
    list_n1.append(cn1)
    list_p2.append(cp2)
    list_n2.append(cn2)
    list_p3.append(cp3)
    list_n3.append(cn3)
    list_p4.append(cp4)
    list_n4.append(cn4)
    if flag==1:
        break


temp=0
for i in range(len(oct_id)+1):
    j=i
    temp+=1
    if i==len(oct_id):
        df.loc[i,'OctantID']='Verified'
        df.loc[i,"1"]=list_p1[0]
        df.loc[i,"-1"]=list_n1[0]
        df.loc[i,"2"]=list_p2[0]
        df.loc[i,"-2"]=list_n2[0]
        df.loc[i,"3"]=list_p3[0]
        df.loc[i,"-3"]=list_n3[0]
        df.loc[i,"4"]=list_p4[0]
        df.loc[i,"-4"]=list_n4[0]
    else:
        df.loc[i,'OctantID']=oct_id[i]
        df.loc[i,"1"]=list_p1[j]
        df.loc[i,"-1"]=list_n1[j]
        df.loc[i,"2"]=list_p2[j]
        df.loc[i,"-2"]=list_n2[j]
        df.loc[i,"3"]=list_p3[j]
        df.loc[i,"-3"]=list_n3[j]
        df.loc[i,"4"]=list_p4[j]
        df.loc[i,"-4"]=list_n4[j]

temp=temp+2
df.loc[temp,'OctantID']='overall Transition Count'
temp+=1
df.loc[temp,'1']='To'
df.loc[temp+2,'']='From'
temp+=1

#formation of 8*8 matrix

array=[]
row,col=8,8
for i in range(row):
    clmn=[]
    for j in range(col):
        clmn.append(0)
    array.append(clmn)

dict={1:0,-1:1,2:2,-2:3,3:4,-3:5,4:6,-4:7}
for i in range(len(df['octant'])-1):
    array[dict[oct[i]]][dict[oct[i+1]]]=array[dict[oct[i]]][dict[oct[i+1]]]+1

hd=["count","+1","-1","+2","-2","+3","-3","+4","-4"]
hd1=["+1","-1","+2","-2","+3","-3","+4","-4"]

for i in range(len(hd)):
    df.loc[temp,'OctantID']=hd[i]
    if i==0:
        df.loc[temp,"1"]=hd1[0]
        df.loc[temp,"-1"]=hd1[1]
        df.loc[temp,"2"]=hd1[2]
        df.loc[temp,"-2"]=hd1[3]
        df.loc[temp,"3"]=hd1[4]
        df.loc[temp,"-3"]=hd1[5]
        df.loc[temp,"4"]=hd1[6]
        df.loc[temp,"-4"]=hd1[7]

    if i!=0:
        
        df.loc[temp,"1"]=array[i-1][0]
        df.loc[temp,"-1"]=array[i-1][1]
        df.loc[temp,"2"]=array[i-1][2]
        df.loc[temp,"-2"]=array[i-1][3]
        df.loc[temp,"3"]=array[i-1][4]
        df.loc[temp,"-3"]=array[i-1][5]
        df.loc[temp,"4"]=array[i-1][6]
        df.loc[temp,"-4"]=array[i-1][7]
    temp+=1



import openpyxl 
from openpyxl.styles import PatternFill
wb = openpyxl.load_workbook("‌input_trial.xlsx")
ws = wb['Sheet1']
fill_cell4 = PatternFill(patternType='solid', fgColor='FCBA03')
ws['L3'].fill = fill_cell4
wb.save("output_octant_transition_identify.xlsx")
flag=0
count=0
while True:
    cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
    t_oct=[]
    num = count + mod
    nI = count + 1
    if (count + mod) > 29745:
        num = 29745
    if count == 0:
        nI = 1
    temp+=2
    df.loc[temp,"OctantID"]="MOD Transition Count"
    temp+=1
    df.loc[temp,"OctantID"]=f"{nI-1} - {num-1}"
    df.loc[temp,"1"]="To"
    df.loc[temp+2,""]="From"
    temp+=1
    oct_id.append(f"{nI-1} - {num-1}")

    for i in range(count,count+mod):
        if i >= len(df['U']):
            flag = 1
            break
        if oct[i] == 1:
            t_oct.append(1)
            cp1+= 1
        if oct[i] == -1:
            t_oct.append(-1)
            cn1 += 1
        if oct[i] == 2:
            t_oct.append(2)
            cp2 += 1
        if oct[i] == -2:
            t_oct.append(-2)
            cn2 += 1
        if oct[i] == 3:
            t_oct.append(3)
            cp3 += 1
        if oct[i] == -3:
            t_oct.append(-3)
            cn3 += 1
        if oct[i] == 4:
            t_oct.append(4)
            cp4 += 1
        if oct[i] == -4:
            t_oct.append(-4)
            cn4+= 1
        count += 1

    arr1=[[0 for i in range(8)] for i in range(8)]

    for i in range(len(t_oct)-1):
        arr1[dict[t_oct[i]]][dict[t_oct[i+1]]]+=1
    
    header=["count","+1","-1","+2","-2","+3","-3","+4","-4"]
    hd1=["+1","-1","+2","-2","+3","-3","+4","-4"]
    for i in range(len(header)):
        df.loc[temp,"OctantID"]=header[i]
        if i==0:
            df.loc[temp,"1"]=hd1[0]
            df.loc[temp,"-1"]=hd1[1]
            df.loc[temp,"2"]=hd1[2]
            df.loc[temp,"-2"]=hd1[3]
            df.loc[temp,"3"]=hd1[4]
            df.loc[temp,"-3"]=hd1[5]
            df.loc[temp,"4"]=hd1[6]
            df.loc[temp,"-4"]=hd1[7]
        if i!=0:
            df.loc[temp,"1"]=arr1[i-1][0]
            df.loc[temp,"-1"]=arr1[i-1][1]
            df.loc[temp,"2"]=arr1[i-1][2]
            df.loc[temp,"-2"]=arr1[i-1][3]
            df.loc[temp,"3"]=arr1[i-1][4]
            df.loc[temp,"-3"]=arr1[i-1][5]
            df.loc[temp,"4"]=arr1[i-1][6]
            df.loc[temp,"-4"]=arr1[i-1][7]
        temp+=1
    
    if flag==1:
        break    







df.to_excel("output_octant_transition_identify.xlsx")