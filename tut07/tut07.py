
# import some library 
import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import load_workbook


from openpyxl.styles.borders import Border, Side  
from openpyxl.styles import PatternFill

start_time = datetime.now()

data = os.listdir('input')
df = pd.DataFrame()
for file in data:
    path = 'input' + '/' + file
    wb = load_workbook(path)
    sheet1 = wb.active
    n = sheet1.max_row

    temp = pd.read_excel(path)
    df = df.append(temp, ignore_index=True)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    
    

    
    l1, l2, l3, time = ([] for i in range(4))

    
    sheet1.cell(row=1, column=8).value = "U'=U-U avg"
    sheet1.cell(row=1, column=9).value = "V'=V-V avg"
    sheet1.cell(row=1, column=10).value = "W=W-W avg"
    sheet1.cell(row=1, column=5).value = "U avg"
    sheet1.cell(row=1, column=6).value = "V avg"
    sheet1.cell(row=1, column=7).value = "W avg"

    Uavg = df['U'].mean()
    Vavg = df['V'].mean()
    Wavg = df['W'].mean()

    

    sheet1['E2'] = round(Uavg, 3)
    sheet1['F2'] = round(Vavg, 3)
    sheet1['G2'] = round(Wavg, 3)

   
    i = 2
    while (i <= n):
        sheet1.cell(row=i, column=8).value = round(sheet1.cell(
            row=i, column=2).value - Uavg, 3)
        l1.append(sheet1.cell(row=i, column=2).value - Uavg)
        sheet1.cell(row=i, column=9).value = round(sheet1.cell(
            row=i, column=3).value - Vavg, 3)
        l2.append(sheet1.cell(row=i, column=3).value - Vavg)
        sheet1.cell(row=i, column=10).value = round(sheet1.cell(
            row=i, column=4).value - Wavg, 3)
        l3.append(sheet1.cell(row=i, column=4).value - Wavg)
        time.append(sheet1.cell(row=i, column=1).value)
        i = i+1

    sheet1.cell(row=1, column=11).value = "Octant"
    o_c_t_list = []
    len5 = 0

    
    i = 0
    o_c1 = o_c2 = o_c3 = o_c4 = o_cneg1 = o_cneg2 = o_cneg3 = o_cneg4 = 0
    while (i < len(l1)):
        if (l1[i] > 0 and l2[i] > 0):
            if (l3[i] > 0):
                o_c1 += 1
                o_c_t_list.insert(len5, +1)
                len5 += 1
            else:
                o_cneg1 += 1
                o_c_t_list.insert(len5, -1)
                len5 += 1
        
        
        elif (l1[i] < 0 and l2[i] > 0):
            if (l3[i] > 0):
                o_c2 += 1
                o_c_t_list.insert(len5, +2)
                len5 += 1
            else:
                o_cneg2 += 1
                o_c_t_list.insert(len5, -2)
                len5 += 1
        elif (l1[i] > 0 and l2[i] < 0):
            if (l3[i] > 0):
                o_c4 += 1
                o_c_t_list.insert(len5, +4)
                len5 += 1
            else:
                o_cneg4 += 1
                o_c_t_list.insert(len5, -4)
                len5 += 1
        elif (l1[i] < 0 and l2[i] < 0):
            if (l3[i] > 0):
                o_c3 += 1
                o_c_t_list.insert(len5, +3)
                len5 += 1
            else:
                o_cneg3 += 1
                o_c_t_list.insert(len5, -3)
                len5 += 1
        i = i+1

    
    i = 2
    while (i <= len(o_c_t_list)+1):
        sheet1.cell(row=i, column=11).value = o_c_t_list[i-2]
        i = i+1

    o_c_t_id = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                              "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
    
    count_Rank1_ModValue = []
    

    def octant_range_names(mod):
        j = k = 0
        interval = mod

        while (j < times+1):






            ct_1 = ct_2 = ct_3 = ct_4 = ct_neg1 = ct_neg2 = ct_neg3 = ct_neg4 = 0
            while (k < mod and k < len(l1)):
                if (l1[k] > 0 and l2[k] > 0):
                    if (l3[k] > 0):
                        ct_1 += 1
                    else:
                        ct_neg1 += 1
                
                elif (l1[k] < 0 and l2[k] < 0):
                    if (l3[k] > 0):
                        ct_3 += 1
                    else:
                        ct_neg3 += 1
                elif (l1[k] < 0 and l2[k] > 0):
                    if (l3[k] > 0):
                        ct_2 += 1
                    else:
                        ct_neg2 += 1
                elif (l1[k] > 0 and l2[k] < 0):
                    if (l3[k] > 0):
                        ct_4 += 1
                    else:
                        ct_neg4 += 1
                k = k+1

            mod = mod+interval
            
            
            
            val_1.append(ct_neg1)
            val2.append(ct_2)
            val_2.append(ct_neg2)
            val3.append(ct_3)
            val_3.append(ct_neg3)
            val4.append(ct_4)
            val1.append(ct_1)
            val_4.append(ct_neg4)
            j = j+1

            mod_list = [ct_1, ct_neg1, ct_2, ct_neg2,
                        ct_3, ct_neg3, ct_4, ct_neg4]
            s_o_t_e_d_lst = [ct_1, ct_neg1, ct_2,
                           ct_neg2, ct_3, ct_neg3, ct_4, ct_neg4]
            s_o_t_e_d_lst.sort(reverse=True)
            
            sheet1.cell(row=j+4, column=23).value = s_o_t_e_d_lst.index(ct_1)+1
            sheet1.cell(
                row=j+4, column=24).value = s_o_t_e_d_lst.index(ct_neg1)+1
            sheet1.cell(row=j+4, column=25).value = s_o_t_e_d_lst.index(ct_2)+1
            sheet1.cell(
                row=j+4, column=26).value = s_o_t_e_d_lst.index(ct_neg2)+1
            sheet1.cell(row=j+4, column=27).value = s_o_t_e_d_lst.index(ct_3)+1
            sheet1.cell(
                row=j+4, column=28).value = s_o_t_e_d_lst.index(ct_neg3)+1
            sheet1.cell(row=j+4, column=29).value = s_o_t_e_d_lst.index(ct_4)+1
            sheet1.cell(
                row=j+4, column=30).value = s_o_t_e_d_lst.index(ct_neg4)+1

            for i in range(23, 31):
                sheet1.cell(row=j+4, column=i).border = thin_border
                if sheet1.cell(row=j+4, column=i).value == 1:
                    sheet1.cell(row=j+4, column=i).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")

            
            if (s_o_t_e_d_lst.index(ct_1)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = 1
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["1"]
            elif (s_o_t_e_d_lst.index(ct_neg1)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = -1
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["-1"]
            
            
            elif (s_o_t_e_d_lst.index(ct_neg2)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = -2
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["-2"]
            elif (s_o_t_e_d_lst.index(ct_3)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = 3
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["3"]
            elif (s_o_t_e_d_lst.index(ct_neg3)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = -3
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["-3"]
            elif (s_o_t_e_d_lst.index(ct_4)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = 4
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["4"]
            elif (s_o_t_e_d_lst.index(ct_2)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = 2
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["2"]
            elif (s_o_t_e_d_lst.index(ct_neg4)+1 == 1):
                sheet1.cell(row=j+4, column=31).value = -4
                sheet1.cell(
                    row=j+4, column=32).value = o_c_t_id["-4"]

            sheet1.cell(row=j+4, column=31).border = thin_border
            sheet1.cell(row=j+4, column=32).border = thin_border

            count_Rank1_ModValue.append(sheet1.cell(row=j+4, column=31).value)

    
    mod = 5000
    lst_tmp = []
    lst_tmp.insert(0, "Overall Count")
    s_t_r = "Mod "
    s_t_r += str(mod)
    lst_tmp.insert(1, s_t_r)
    num = n
    times = int(num/mod)
    modulo_val = num % mod
    len6 = 2
    for i in range(times):
        s_t_r = str(mod*i)
        s_t_r += "-"
        s_t_r += str(mod*(i+1)-1)
        lst_tmp.insert(len6, s_t_r)
        len6 += 1
    if (modulo_val):
        s_t_r = str(num-modulo_val)
        s_t_r += "-"
        s_t_r += str(num)
        lst_tmp.insert(len6, s_t_r)
        len6 += 1
    sheet1.cell(row=3, column=13).value = "Mod" + str(mod)
    sheet1.cell(row=2, column=14).value = "octant ID"
    sheet1.cell(row=2, column=14).border = thin_border
    i = 3
    while (i <= len(lst_tmp)+2):
        sheet1.cell(row=i, column=14).value = lst_tmp[i-3]
        sheet1.cell(row=i, column=14).border = thin_border
        i = i+1

    lst_occt = ["1", "-1", "2", "-2", "3", "-3", "4", "-4"]
    r_k_lst = ["Rank Octant 1", "Rank Octant -1", "Rank Octant 2", "Rank Octant -2",
                 "Rank Octant 3", "Rank Octant -3", "Rank Octant 4", "Rank Octant -4"]

    
    i = 0
    j = 14
    while (i < len(lst_occt)):
        sheet1.cell(row=2, column=j+1).value = lst_occt[i]
        sheet1.cell(row=2, column=j+1).border = thin_border
        sheet1.cell(row=2, column=j+9).value = r_k_lst[i]
        sheet1.cell(row=2, column=j+9).border = thin_border
        j = j+1
        i = i+1
    sheet1.cell(row=2, column=31).value = "Rank1 Octant ID"
    sheet1.cell(row=2, column=31).border = thin_border
    sheet1.cell(row=2, column=32).value = "Rank1 Octant Name"
    sheet1.cell(row=2, column=32).border = thin_border

    
    val1, val_1, val2, val_2, val3, val_3, val4, val_4 = ([] for i in range(8))

    
    
    
    val2.insert(0, o_c2)
    val_2.insert(0, o_cneg2)
    val_4.insert(0, o_cneg4)
    val3.insert(0, o_c3)
    val_3.insert(0, o_cneg3)
    val4.insert(0, o_c4)
    val1.insert(0, o_c1)
    val_1.insert(0, o_cneg1)
    
    
    
    val_1.insert(1, None)
    val2.insert(1, None)
    val3.insert(1, None)
    val_3.insert(1, None)
    val4.insert(1, None)
    val1.insert(1, None)
    val_2.insert(1, None)
    val_4.insert(1, None)

    
    octant_range_names(mod)

    
    i = 3
    while (i <= len(val1)+2):
        
        sheet1.cell(row=i, column=19).value = val3[i-3]
        
        sheet1.cell(row=i, column=17).value = val2[i-3]
        sheet1.cell(row=i, column=18).value = val_2[i-3]
        sheet1.cell(row=i, column=20).value = val_3[i-3]
        sheet1.cell(row=i, column=21).value = val4[i-3]
        sheet1.cell(row=i, column=15).value = val1[i-3]
        sheet1.cell(row=i, column=16).value = val_1[i-3]
        sheet1.cell(row=i, column=22).value = val_4[i-3]
        for j in range(15, 23):
            sheet1.cell(row=i, column=j).border = thin_border
        i = i+1

    sheet1.cell(row=i+3, column=29).value = "Octant ID"
    sheet1.cell(row=i+3, column=30).value = "Octant Name"
    sheet1.cell(row=i+3, column=31).value = "Count of Rank 1 Mod Values"
    for col in range(29, 32):
        sheet1.cell(row=i+3, column=col).border = thin_border
    k = 0
    j = i+4
    while (k < len(lst_occt)):
        sheet1.cell(row=j, column=29).value = lst_occt[k]
        sheet1.cell(row=j, column=29).border = thin_border
        sheet1.cell(
            row=j, column=30).value = o_c_t_id[lst_occt[k]]
        sheet1.cell(row=j, column=30).border = thin_border
        sheet1.cell(row=j, column=31).value = count_Rank1_ModValue.count(
            int(lst_occt[k]))
        sheet1.cell(row=j, column=31).border = thin_border
        j = j+1
        k = k+1

    overall_list = [val1[0], val_1[0], val2[0],
                    val_2[0], val3[0], val_3[0], val4[0], val_4[0]]
    overall_s_o_t_e_d_lst = [val1[0], val_1[0], val2[0],
                           val_2[0], val3[0], val_3[0], val4[0], val_4[0]]
    overall_s_o_t_e_d_lst.sort(reverse=True)
    
    sheet1.cell(row=3, column=27).value = overall_s_o_t_e_d_lst.index(val3[0])+1
    sheet1.cell(row=3, column=28).value = overall_s_o_t_e_d_lst.index(val_3[0])+1
    sheet1.cell(row=3, column=29).value = overall_s_o_t_e_d_lst.index(val4[0])+1
    sheet1.cell(row=3, column=23).value = overall_s_o_t_e_d_lst.index(val1[0])+1
    sheet1.cell(row=3, column=24).value = overall_s_o_t_e_d_lst.index(val_1[0])+1
    sheet1.cell(row=3, column=25).value = overall_s_o_t_e_d_lst.index(val2[0])+1
    sheet1.cell(row=3, column=26).value = overall_s_o_t_e_d_lst.index(val_2[0])+1
    sheet1.cell(row=3, column=30).value = overall_s_o_t_e_d_lst.index(val_4[0])+1
    for i in range(23, 31):
        sheet1.cell(row=3, column=i).border = thin_border
        if sheet1.cell(row=3, column=i).value == 1:
            sheet1.cell(row=3, column=i).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")  # to fill color to the cell

    
    if (overall_s_o_t_e_d_lst.index(val1[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = 1
        sheet1.cell(row=3, column=32).value = o_c_t_id["1"]
    
    
    elif (overall_s_o_t_e_d_lst.index(val_3[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = -3
        sheet1.cell(row=3, column=32).value = o_c_t_id["-3"]
    elif (overall_s_o_t_e_d_lst.index(val2[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = 2
        sheet1.cell(row=3, column=32).value = o_c_t_id["2"]
    elif (overall_s_o_t_e_d_lst.index(val4[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = 4
        sheet1.cell(row=3, column=32).value = o_c_t_id["4"]
    elif (overall_s_o_t_e_d_lst.index(val_2[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = -2
        sheet1.cell(row=3, column=32).value = o_c_t_id["-2"]
    
    
    elif (overall_s_o_t_e_d_lst.index(val3[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = 3
        sheet1.cell(row=3, column=32).value = o_c_t_id["3"]
    
    elif (overall_s_o_t_e_d_lst.index(val_1[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = -1
        sheet1.cell(row=3, column=32).value = o_c_t_id["-1"]
    elif (overall_s_o_t_e_d_lst.index(val_4[0])+1 == 1):
        sheet1.cell(row=3, column=31).value = -4
        sheet1.cell(row=3, column=32).value = o_c_t_id["-4"]

    sheet1.cell(row=3, column=31).border = thin_border
    sheet1.cell(row=3, column=32).border = thin_border

    sheet1.cell(row=1, column=35).value = "Overall Transition Count"
    sheet1.cell(row=2, column=36).value = "To"
    

    lst_occt = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    j = 36
    i = 0
    while (j < 44 and i < len(lst_occt)):
        sheet1.cell(row=3, column=j).value = lst_occt[i]
        sheet1.cell(row=3, column=j).border = thin_border
        j = j+1
        i = i+1

    sheet1.cell(row=3, column=35).value = "Count"
    sheet1.cell(row=3, column=35).border = thin_border
    sheet1.cell(row=4, column=34).value = "From"

    k = 4
    i = 0
    while (k < 12 and i < len(lst_occt)):
        sheet1.cell(row=k, column=35).value = lst_occt[i]
        sheet1.cell(row=k, column=35).border = thin_border
        k = k+1
        i = i+1

    
    
    
    r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
    r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0
    r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
    r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
    r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
    r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
    r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0
    r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
    
    i = 0
    while (i < len(o_c_t_list)-1):
        if (o_c_t_list[i] == 1):
            if (o_c_t_list[i+1] == 1):
                r1c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r1c2 += 1
            elif (o_c_t_list[i+1] == 2):
                r1c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r1c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r1c5 += 1
            elif (o_c_t_list[i+1] == -3):
                r1c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r1c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r1c8 += 1
            i += 1
        
        elif (o_c_t_list[i] == 2):
            if (o_c_t_list[i+1] == 1):
                r3c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r3c2 += 1
            elif (o_c_t_list[i+1] == 2):
                r3c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r3c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r3c5 += 1
            elif (o_c_t_list[i+1] == -3):
                r3c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r3c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r3c8 += 1
            i += 1
        
        elif (o_c_t_list[i] == -1):
            if (o_c_t_list[i+1] == 1):
                r2c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r2c2 += 1
            elif (o_c_t_list[i+1] == 2):
                r2c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r2c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r2c5 += 1
            elif (o_c_t_list[i+1] == -3):
                r2c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r2c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r2c8 += 1
            i += 1
        elif (o_c_t_list[i] == -3):
            if (o_c_t_list[i+1] == 1):
                r6c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r6c2 += 1
            elif (o_c_t_list[i+1] == 2):
                r6c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r6c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r6c5 += 1
            elif (o_c_t_list[i+1] == -3):
                r6c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r6c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r6c8 += 1
            i += 1
        elif (o_c_t_list[i] == -2):
            if (o_c_t_list[i+1] == 1):
                r4c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r4c2 += 1
            
            elif (o_c_t_list[i+1] == -3):
                r4c6 += 1
            elif (o_c_t_list[i+1] == 2):
                r4c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r4c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r4c5 += 1
            elif (o_c_t_list[i+1] == 4):
                r4c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r4c8 += 1
            i += 1
        elif (o_c_t_list[i] == -4):
            if (o_c_t_list[i+1] == 1):
                r8c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r8c2 += 1
            
            elif (o_c_t_list[i+1] == -3):
                r8c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r8c7 += 1
            elif (o_c_t_list[i+1] == 2):
                r8c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r8c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r8c5 += 1
            elif (o_c_t_list[i+1] == -4):
                r8c8 += 1
            i += 1
        elif (o_c_t_list[i] == 3):
            if (o_c_t_list[i+1] == 1):
                r5c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r5c2 += 1
            elif (o_c_t_list[i+1] == 2):
                r5c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r5c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r5c5 += 1
            elif (o_c_t_list[i+1] == -3):
                r5c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r5c7 += 1
            elif (o_c_t_list[i+1] == -4):
                r5c8 += 1
            
            i += 1
        elif (o_c_t_list[i] == 4):
            if (o_c_t_list[i+1] == 1):
                r7c1 += 1
            elif (o_c_t_list[i+1] == -1):
                r7c2 += 1
            
            elif (o_c_t_list[i+1] == -3):
                r7c6 += 1
            elif (o_c_t_list[i+1] == 4):
                r7c7 += 1
            elif (o_c_t_list[i+1] == 2):
                r7c3 += 1
            elif (o_c_t_list[i+1] == -2):
                r7c4 += 1
            elif (o_c_t_list[i+1] == 3):
                r7c5 += 1
            elif (o_c_t_list[i+1] == -4):
                r7c8 += 1
            i += 1
        
   
    
    row_transition3 = [r1c3, r2c3, r3c3, r4c3, r5c3, r6c3, r7c3, r8c3]
    row_transition4 = [r1c4, r2c4, r3c4, r4c4, r5c4, r6c4, r7c4, r8c4]
    row_transition5 = [r1c5, r2c5, r3c5, r4c5, r5c5, r6c5, r7c5, r8c5]
    row_transition6 = [r1c6, r2c6, r3c6, r4c6, r5c6, r6c6, r7c6, r8c6]
    row_transition7 = [r1c7, r2c7, r3c7, r4c7, r5c7, r6c7, r7c7, r8c7]
    row_transition8 = [r1c8, r2c8, r3c8, r4c8, r5c8, r6c8, r7c8, r8c8]
    row_transition1 = [r1c1, r2c1, r3c1, r4c1, r5c1, r6c1, r7c1, r8c1]
    row_transition2 = [r1c2, r2c2, r3c2, r4c2, r5c2, r6c2, r7c2, r8c2]

    
    i = 0
    j = 4
    while (j <= 11 and i < len(row_transition1)):
        sheet1.cell(row=j, column=36).value = row_transition1[i]
        if i == 0:
            sheet1.cell(row=j, column=36).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=37).value = row_transition2[i]
        
        
        if i == 3:
            sheet1.cell(row=j, column=39).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=40).value = row_transition5[i]
        if i == 1:
            sheet1.cell(row=j, column=37).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=38).value = row_transition3[i]
        if i == 2:
            sheet1.cell(row=j, column=38).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=39).value = row_transition4[i]
        
        
        if i == 6:
            sheet1.cell(row=j, column=42).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=43).value = row_transition8[i]

        if i == 5:
            sheet1.cell(row=j, column=41).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=42).value = row_transition7[i]
        if i == 4:
            sheet1.cell(row=j, column=40).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        sheet1.cell(row=j, column=41).value = row_transition6[i]
        if i == 7:
            sheet1.cell(row=j, column=43).fill = PatternFill(
                start_color='FFD970', end_color='FFD970', fill_type="solid")
        for col in range(36, 44):
            sheet1.cell(row=j, column=col).border = thin_border
        i = i+1
        j = j+1

    

    def transition_identification(mod):
        x = 0
        i = 0
        p = 1
        interval = mod

        while (x < n/interval):
            l = 36
            j = 0
            while (l < 44 and j < len(lst_occt)):
                sheet1.cell(row=p*13+2, column=l).value = lst_occt[j]
                sheet1.cell(row=p*13+2, column=l).border = thin_border
                l = l+1
                j = j+1
            sheet1.cell(row=p*13,
                        column=35).value = "Mod Transition Count"
            sheet1.cell(row=1+p*13,
                        column=35).value = str(interval*x)+"-"+str(interval*(x+1)-1)
            sheet1.cell(row=1+p*13, column=36).value = "To"
            sheet1.cell(row=2+p*13, column=35).value = "Count"
            sheet1.cell(row=2+p*13, column=35).border = thin_border
            sheet1.cell(row=3+p*13, column=34).value = "From"

            m = 3+p*13
            q = 0
            while (m < 11+p*13 and q < len(lst_occt)):
                sheet1.cell(row=m, column=35).value = lst_occt[q]
                sheet1.cell(row=m, column=35).border = thin_border
                m = m+1
                q = q+1
            
            
            r5c1 = r5c2 = r5c3 = r5c4 = r5c5 = r5c6 = r5c7 = r5c8 = 0
            r6c1 = r6c2 = r6c3 = r6c4 = r6c5 = r6c6 = r6c7 = r6c8 = 0
            r7c1 = r7c2 = r7c3 = r7c4 = r7c5 = r7c6 = r7c7 = r7c8 = 0
            r8c1 = r8c2 = r8c3 = r8c4 = r8c5 = r8c6 = r8c7 = r8c8 = 0
            r1c1 = r1c2 = r1c3 = r1c4 = r1c5 = r1c6 = r1c7 = r1c8 = 0
            r2c1 = r2c2 = r2c3 = r2c4 = r2c5 = r2c6 = r2c7 = r2c8 = 0
            r3c1 = r3c2 = r3c3 = r3c4 = r3c5 = r3c6 = r3c7 = r3c8 = 0
            r4c1 = r4c2 = r4c3 = r4c4 = r4c5 = r4c6 = r4c7 = r4c8 = 0

            # data_frame["octant ID"] = L_TMp
#     # inserting all columns of their octant value
#     data_frame["1"] = val1
#     data_frame["-1"] = vaneg1
#     data_frame["2"] = val2
#     data_frame["-2"] = vaneg2
#     data_frame["3"] = val3
#     data_frame["-3"] = vaneg3
#     data_frame["4"] = val4
#     data_frame["-4"] = vaneg4

            while (i < mod and i < len(o_c_t_list)-1):
                if (o_c_t_list[i] == 1):
                    if (o_c_t_list[i+1] == 1):
                        r1c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r1c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r1c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r1c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r1c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r1c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r1c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r1c8 += 1

                

                elif (o_c_t_list[i] == 2):
                    if (o_c_t_list[i+1] == 1):
                        r3c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r3c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r3c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r3c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r3c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r3c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r3c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r3c8 += 1

                
                elif (o_c_t_list[i] == 3):
                    if (o_c_t_list[i+1] == 1):
                        r5c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r5c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r5c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r5c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r5c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r5c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r5c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r5c8 += 1
                elif (o_c_t_list[i] == -1):
                    if (o_c_t_list[i+1] == 1):
                        r2c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r2c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r2c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r2c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r2c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r2c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r2c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r2c8 += 1
                elif (o_c_t_list[i] == -2):
                    if (o_c_t_list[i+1] == 1):
                        r4c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r4c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r4c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r4c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r4c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r4c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r4c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r4c8 += 1


                elif (o_c_t_list[i] == 4):
                    if (o_c_t_list[i+1] == 1):
                        r7c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r7c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r7c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r7c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r7c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r7c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r7c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r7c8 += 1


                elif (o_c_t_list[i] == -3):
                    if (o_c_t_list[i+1] == 1):
                        r6c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r6c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r6c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r6c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r6c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r6c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r6c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r6c8 += 1
                elif (o_c_t_list[i] == -4):
                    if (o_c_t_list[i+1] == 1):
                        r8c1 += 1
                    elif (o_c_t_list[i+1] == -1):
                        r8c2 += 1
                    elif (o_c_t_list[i+1] == 2):
                        r8c3 += 1
                    elif (o_c_t_list[i+1] == -2):
                        r8c4 += 1
                    elif (o_c_t_list[i+1] == 3):
                        r8c5 += 1
                    elif (o_c_t_list[i+1] == -3):
                        r8c6 += 1
                    elif (o_c_t_list[i+1] == 4):
                        r8c7 += 1
                    elif (o_c_t_list[i+1] == -4):
                        r8c8 += 1
                i += 1
            
            
            row_transition5 = [r1c5, r2c5, r3c5, r4c5, r5c5, r6c5, r7c5, r8c5]
            
            row_transition2 = [r1c2, r2c2, r3c2, r4c2, r5c2, r6c2, r7c2, r8c2]
            row_transition3 = [r1c3, r2c3, r3c3, r4c3, r5c3, r6c3, r7c3, r8c3]
            row_transition6 = [r1c6, r2c6, r3c6, r4c6, r5c6, r6c6, r7c6, r8c6]
            row_transition7 = [r1c7, r2c7, r3c7, r4c7, r5c7, r6c7, r7c7, r8c7]
            row_transition8 = [r1c8, r2c8, r3c8, r4c8, r5c8, r6c8, r7c8, r8c8]
            row_transition1 = [r1c1, r2c1, r3c1, r4c1, r5c1, r6c1, r7c1, r8c1]
            row_transition4 = [r1c4, r2c4, r3c4, r4c4, r5c4, r6c4, r7c4, r8c4]
            
            y = 0
            z = p*13+3
            while (z <= 10+p*13 and y < len(row_transition1)):
                sheet1.cell(row=z, column=36).value = row_transition1[y]
                
                if y == 1:
                    sheet1.cell(row=z, column=37).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=38).value = row_transition3[y]
                if y == 0:
                    sheet1.cell(row=z, column=36).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=37).value = row_transition2[y]
                if y == 3:
                    sheet1.cell(row=z, column=39).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=40).value = row_transition5[y]
                if y == 2:
                    sheet1.cell(row=z, column=38).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=39).value = row_transition4[y]
                if y == 5:
                    sheet1.cell(row=z, column=41).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=42).value = row_transition7[y]

                if y == 7:
                    sheet1.cell(row=z, column=43).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")

                if y == 4:
                    sheet1.cell(row=z, column=40).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=41).value = row_transition6[y]

                if y == 6:
                    sheet1.cell(row=z, column=42).fill = PatternFill(
                        start_color='FFD970', end_color='FFD970', fill_type="solid")
                sheet1.cell(row=z, column=43).value = row_transition8[y]
                
                for col in range(36, 44):
                    sheet1.cell(row=z, column=col).border = thin_border
                y = y+1
                z = z+1
            mod = mod+interval
            p = p+1
            x = x+1

   # oct_id=['Overall count',f"mod-{mod}"]
# flag=count=0
# while True:
#     cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
#     num=count+mod
#     nI=1+count
#     if(count+mod)>29745:
#         num=29745
#     if(count==0):
#         nI=0
#     oct_id.append(f"{nI-1}-{num-1}") 
    transition_identification(mod)

    def octant_longest_subsequence_count():
        
        count1 = count_1 = count2 = count_2 = count3 = count_3 = count4 = count_4 = 1

        for i in range(len(o_c_t_list)-1):
            
            if (o_c_t_list[i] == 1):
                if (o_c_t_list[i+1] == 1):
                    count1 += 1
                else:
                    list1.append(count1)
                    count1 = 1
            
            elif (o_c_t_list[i] == 2):
                if (o_c_t_list[i+1] == 2):
                    count2 += 1
                else:
                    list2.append(count2)
                    count2 = 1

            elif (o_c_t_list[i] == -1):
                if (o_c_t_list[i+1] == -1):
                    count_1 += 1
                else:
                    l_s_t1.append(count_1)
                    count_1 = 1
            
            elif (o_c_t_list[i] == 3):
                if (o_c_t_list[i+1] == 3):
                    count3 += 1
                else:
                    list3.append(count3)
                    count3 = 1

            elif (o_c_t_list[i] == -2):
                if (o_c_t_list[i+1] == -2):
                    count_2 += 1
                else:
                    l_s_t2.append(count_2)
                    count_2 = 1
            
            elif (o_c_t_list[i] == 4):
                if (o_c_t_list[i+1] == 4):
                    count4 += 1
                else:
                    list4.append(count4)
                    count4 = 1

            elif (o_c_t_list[i] == -3):
                if (o_c_t_list[i+1] == -3):
                    count_3 += 1
                else:
                    l_s_t3.append(count_3)
                    count_3 = 1
            
            elif (o_c_t_list[i] == -4):
                if (o_c_t_list[i+1] == -4):
                    count_4 += 1
                else:
                    l_s_t4.append(count_4)
                    count_4 = 1

    
    sheet1.cell(row=3, column=46).border = thin_border
    sheet1.cell(row=3, column=47).value = "Count"
    sheet1.cell(row=3, column=47).border = thin_border
    sheet1.cell(row=3, column=45).value = "Octant"
    sheet1.cell(row=3, column=45).border = thin_border
    sheet1.cell(row=3, column=46).value = "Longest Subsequence Length"

    lst_occt = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    i = 0
    while (i < len(lst_occt)):
        sheet1.cell(row=i+4, column=45).value = lst_occt[i]
        sheet1.cell(row=i+4, column=45).border = thin_border
        i = i+1

    
    
    l_s_t3 = []
    list4 = []
    l_s_t4 = []
    list1 = []
    l_s_t1 = []
    list2 = []
    l_s_t2 = []
    list3 = []

    
    octant_longest_subsequence_count()

    

    
    sheet1.cell(row=9, column=46).value = max(l_s_t3)
    sheet1.cell(row=10, column=46).value = max(list4)
    sheet1.cell(row=11, column=46).value = max(l_s_t4)
    sheet1.cell(row=4, column=46).value = max(list1)
    sheet1.cell(row=5, column=46).value = max(l_s_t1)
    sheet1.cell(row=6, column=46).value = max(list2)
    sheet1.cell(row=7, column=46).value = max(l_s_t2)
    sheet1.cell(row=8, column=46).value = max(list3)



# dic={1:ocp1,-1:ocn1,2:ocp2,-2:ocn2,3:ocp3,-3:ocn3,4:ocp4,-4:ocn4}
# for i in dic:
#     if(dic[i]==list2[0]):
#         list_rank1.append(i)

# list_rank1.append(' ')

# k=0
# for i in list1:
#     idx=list2.index(i)
#     Sheet1.cell(row=2,column=22+k).value=idx+1
#     k+=1
# intv=1

    for r in range(4, 12):
        sheet1.cell(row=r, column=46).border = thin_border

    
    
    sheet1.cell(row=9, column=47).value = l_s_t3.count(max(l_s_t3))
    sheet1.cell(row=10, column=47).value = list4.count(max(list4))
    sheet1.cell(row=11, column=47).value = l_s_t4.count(max(l_s_t4))
    sheet1.cell(row=4, column=47).value = list1.count(max(list1))
    sheet1.cell(row=5, column=47).value = l_s_t1.count(max(l_s_t1))
    sheet1.cell(row=6, column=47).value = list2.count(max(list2))
    sheet1.cell(row=7, column=47).value = l_s_t2.count(max(l_s_t2))
    sheet1.cell(row=8, column=47).value = list3.count(max(list3))

    for r in range(4, 12):
        sheet1.cell(row=r, column=47).border = thin_border

    def octant_longest_subsequence_count_with_range():
        
        count1 = count_1 = count2 = count_2 = count3 = count_3 = count4 = count_4 = 1

        for i in range(len(o_c_t_list)-1):
            
            if (o_c_t_list[i] == 1):
                if (o_c_t_list[i+1] == 1):
                    count1 += 1
                else:
                    if (count1 == max(list1)):
                        time1.append([time[i-count1+1], time[i]])
                    count1 = 1

            
            elif (o_c_t_list[i] == 2):
                if (o_c_t_list[i+1] == 2):
                    count2 += 1
                else:
                    if (count2 == max(list2)):
                        time2.append([time[i-count2+1], time[i]])
                    count2 = 1

            elif (o_c_t_list[i] == -1):
                if (o_c_t_list[i+1] == -1):
                    count_1 += 1
                else:
                    if (count_1 == max(l_s_t1)):
                        t_m_e1.append([time[i-count_1+1], time[i]])
                    count_1 = 1

            
            elif (o_c_t_list[i] == 3):
                if (o_c_t_list[i+1] == 3):
                    count3 += 1
                else:
                    if (count3 == max(list3)):
                        time3.append([time[i-count3+1], time[i]])
                    count3 = 1

                    
            elif (o_c_t_list[i] == -2):
                if (o_c_t_list[i+1] == -2):
                    count_2 += 1
                else:
                    if (count_2 == max(l_s_t2)):
                        t_m_e2.append([time[i-count_2+1], time[i]])
                    count_2 = 1

            

            elif (o_c_t_list[i] == 4):
                if (o_c_t_list[i+1] == 4):
                    count4 += 1
                else:
                    if (count4 == max(list4)):
                        time4.append([time[i-count4+1], time[i]])
                    count4 = 1

            elif (o_c_t_list[i] == -3):
                if (o_c_t_list[i+1] == -3):
                    count_3 += 1
                else:
                    if (count_3 == max(l_s_t3)):
                        t_m_e3.append([time[i-count_3+1], time[i]])
                    count_3 = 1

            elif (o_c_t_list[i] == -4):
                if (o_c_t_list[i+1] == -4):
                    count_4 += 1
                else:
                    if (count_4 == max(l_s_t4)):
                        t_m_e4.append([time[i-count_4+1], time[i]])
                    count_4 = 1

    
    sheet1.cell(row=3, column=51).value = "Count"
    sheet1.cell(row=3, column=51).border = thin_border
    sheet1.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
    sheet1.cell(row=3, column=49).value = "Octant"
    sheet1.cell(row=3, column=49).border = thin_border
    sheet1.cell(row=3, column=50).value = "Longest Subsequence Length"
    sheet1.cell(row=3, column=50).border = thin_border
    lst_occt = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
    i = 0
    while (i < len(lst_occt)):
        sheet1.cell(row=i+3, column=49).value = lst_occt[i]
        sheet1.cell(row=i+3, column=49).border = thin_border
        i = i+1

    
    time1, t_m_e1, time2, t_m_e2, time3, t_m_e3, time4, t_m_e4 = (
        [] for i in range(8))

    
    octant_longest_subsequence_count_with_range()

    
    

    col_18 = [max(list1), "From", time1[0][0], max(l_s_t1), "From", t_m_e1[0][0], max(list2), "From", time2[0][0], max(l_s_t2), "From", t_m_e2[0][0], max(
        list3), "From", time3[0][0], max(l_s_t3), "From", t_m_e3[0][0], max(list4), "From", time4[0][0], max(l_s_t4), "From", t_m_e4[0][0]]

    
    col_17 = ["+1", "Time", " ", "-1", "Time", " ", "+2", "Time", " ", "-2",
              "Time", " ", "+3", "Time", " ", "-3", "Time", " ", "+4", "Time", " ", "-4", "Time", " "]
    

    col_19 = [list1.count(max(list1)), "To", time1[0][1], l_s_t1.count(max(l_s_t1)), "To", t_m_e1[0][1], list2.count(max(list2)), "To", time2[0][1], l_s_t2.count(max(l_s_t2)), "To", t_m_e2[0][1], list3.count(
        max(list3)), "To", time3[0][1], l_s_t3.count(max(l_s_t3)), "To", t_m_e3[0][1], list4.count(max(list4)), "To", time4[0][1], l_s_t4.count(max(l_s_t4)), "To", t_m_e4[0][1]]

    
    for i in range(3, len(col_17)+3):
        sheet1.cell(row=i, column=49).value = col_17[i-3]
        sheet1.cell(row=i, column=49).border = thin_border
        sheet1.cell(row=i, column=50).value = col_18[i-3]
        sheet1.cell(row=i, column=50).border = thin_border
        sheet1.cell(row=i, column=51).value = col_19[i-3]
        sheet1.cell(row=i, column=51).border = thin_border

    wb.save('output/' + str(file[:-5]) +
            'octant_analysis_mod' + str(mod) + '.xlsx')

# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))




