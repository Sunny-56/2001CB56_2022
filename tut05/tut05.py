#Help https://youtu.be/N6PBd4XdnEw
# def octant_range_names(mod=5000):

    
#     octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}

# ###Code

# from platform import python_version
# ver = python_version()

# if ver == "3.8.10":
#     print("Correct Version Installed")
# else:
#     print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


# mod=5000 
# octant_range_names(mod)

from array import array
from operator import mod
import pandas as pd
df=pd.read_excel('octant_input.xlsx')
# df.drop(df.columns[4:26],axis=1,inplace=True)
mod=5000
avg_u=df['U'].mean()
avg_v=df['V'].mean()
avg_w=df['W'].mean()
df.loc[0,"U avg"]=avg_u
df.loc[0,"V avg"]=avg_v
df.loc[0,"W avg"]=avg_w

di1=[]
di2=[]
di3=[]

for i in df['U']:
    di1.append(i-avg_u)
for i in df['V']:
    di2.append(i-avg_v)
for i in df['W']:
    di3.append(i-avg_w)

df["U' = U - Avg"]=di1
df["V' = V - Avg"]=di2
df["W' = W - Avg"]=di3

oct=[]
ocp1=ocn1=ocp2=ocn2=ocp3=ocn3=ocp4=ocn4=0
for i in range(len(df['U'])):
    if(di1[i]>0 and di2[i]>0):
        if(di3[i]>0):
            oct.append(1)
            ocp1+=1
        else:
            oct.append(-1)
            ocn1+=1
    elif(di1[i]<0 and di2[i]>0):
        if(di3[i]>0):
            oct.append(2)
            ocp2+=1
        else:
            oct.append(-2)
            ocn2+=1

    elif(di1[i]>0 and di2[i]<0):
        if(di3[i]>0):
            oct.append(4)
            ocp4+=1
        else:
            oct.append(-4)
            ocn4+=1
    
    elif(di1[i]<0 and di2[i]<0):
        if(di3[i]>0):
            oct.append(3)
            ocp3+=1
        else:
            oct.append(-3)
            ocn3+=1
    
    

df["octant"]=oct
df.loc[2,'']='User input'


# dt={1:cp1,-1:cn1,2:cp2,-2:cn2,3:cp3,-3:cn3,4:cp4,-4:cn4}
# sorted_dt = {key: value for key, value in sorted(dt.items(), key=lambda item: item[1])}



list_p1=[]
list_n1=[]
list_p2=[]
list_n2=[]
list_p3=[]
list_n3=[]
list_p4=[]
list_n4=[]

list_p1.append(ocp1)
list_p1.append("")
list_n1.append(ocn1)
list_n1.append('')
list_p2.append(ocp2)
list_p2.append('')
list_n2.append(ocn2)
list_n2.append('')
list_p3.append(ocp3)
list_p3.append('')
list_n3.append(ocn3)
list_n3.append('')
list_p4.append(ocp4)
list_p4.append('')
list_n4.append(ocn4)
list_n4.append('')



lst_t1=[]


oct_id=['Overall count',f"mod-{mod}"]
flag=count=0
l=0

import openpyxl 
from openpyxl.styles import PatternFill
wb = openpyxl.load_workbook("output_ranking.xlsx")
Sheet1=wb.active
list_rank1=[]
list1=[ocp1,ocn1,ocp2,ocn2,ocp3,ocn3,ocp4,ocn4]
list2=list1.copy()
list2.sort(reverse=True)


dic={1:ocp1,-1:ocn1,2:ocp2,-2:ocn2,3:ocp3,-3:ocn3,4:ocp4,-4:ocn4}
for i in dic:
    if(dic[i]==list2[0]):
        list_rank1.append(i)

list_rank1.append(' ')

k=0
for i in list1:
    idx=list2.index(i)
    Sheet1.cell(row=2,column=22+k).value=idx+1
    k+=1
intv=1
while True:
    
    cp1=cn1=cp2=cn2=cp3=cn3=cp4=cn4=0
    num=count+mod
    nI=1+count
    if(count+mod)>29745:
        num=29745
    if(count==0):
        nI=0
    oct_id.append(f"{nI-1}-{num-1}")
    

    
    
    for i in range(count,count+mod):
        if(i>=len(df['U'])):
            flag=1
            break
        if(oct[i]==1):
            cp1+=1
        if(oct[i]==-1):
            cn1+=1
        if(oct[i]==2):
            cp2+=1
        if(oct[i]==-2):
            cn2+=1
        if(oct[i]==3):
            cp3+=1
        if(oct[i]==-3):
            cn3+=1
        if(oct[i]==4):
            cp4+=1
        if(oct[i]==-4):
            cn4+=1
        count+=1
    


    

    list_p1.append(cp1)
    list_n1.append(cn1)
    list_p2.append(cp2)
    list_n2.append(cn2)
    list_p3.append(cp3)
    list_n3.append(cn3)
    list_p4.append(cp4)
    list_n4.append(cn4)
    



    lst_t1.append(cp1)
    lst_t1.append(cn1)
    lst_t1.append(cp2)
    lst_t1.append(cn2)
    lst_t1.append(cp3)
    lst_t1.append(cn3)
    lst_t1.append(cp4)
    lst_t1.append(cn4)
    
    lst_T1=lst_t1.copy()
    lst_t1.sort(reverse=True)


    dt={1:cp1,-1:cn1,2:cp2,-2:cn2,3:cp3,-3:cn3,4:cp4,-4:cn4}
    for i in dt:
        if(dt[i]==lst_t1[0]):
            list_rank1.append(i)

    k=0

       
    for i in lst_T1:
        idx=lst_t1.index(i)
        Sheet1.cell(row=4+l,column=22+k).value=idx+1
        k+=1
    if l>=intv:
        break
    l+=1
    lst_t1.clear()


    intv+=1

    if flag==1:
        break


    


temp=0
for i in range(len(oct_id)):
    j=i
    temp+=1
    
    df.loc[i,'OctantID']=oct_id[i]
    df.loc[i,"1"]=list_p1[j]
    df.loc[i,"-1"]=list_n1[j]
    df.loc[i,"2"]=list_p2[j]
    df.loc[i,"-2"]=list_n2[j]
    df.loc[i,"3"]=list_p3[j]
    df.loc[i,"-3"]=list_n3[j]
    df.loc[i,"4"]=list_p4[j]
    df.loc[i,"-4"]=list_n4[j]
    
df.to_excel("output_ranking.xlsx",index=False)



Sheet1.cell(row=1,column=22).value='Rank of 1'
Sheet1.cell(row=1,column=23).value='Rank of -1'
Sheet1.cell(row=1,column=24).value='Rank of 2'
Sheet1.cell(row=1,column=25).value='Rank of -2'
Sheet1.cell(row=1,column=26).value='Rank of 3'
Sheet1.cell(row=1,column=27).value='Rank of -3'
Sheet1.cell(row=1,column=28).value='Rank of 4'
Sheet1.cell(row=1,column=29).value='Rank of -4'
Sheet1.cell(row=1,column=30).value='Rank1 Octant ID'
Sheet1.cell(row=1,column=31).value='Rank1 Octant Name'


Sheet1['N13']='Octant Id'
Sheet1['O13']='Octant Name'
Sheet1['P13']='Count of Rank1 Mod Values'
Sheet1['N14']='1'
Sheet1['N15']='-1'
Sheet1['N16']='2'
Sheet1['N17']='-2'
Sheet1['N18']='3'
Sheet1['N19']='-3'
Sheet1['N20']='4'
Sheet1['N21']='-4'



Sheet1['O14']='Internal outward interaction'
Sheet1['O15']='External outward interaction'
Sheet1['O16']='External ejection'
Sheet1['O17']='internal ejection'
Sheet1['O18']='External inward interaction'
Sheet1['O19']='Internal Inward interaction'
Sheet1['O20']='Internal sweep'
Sheet1['O21']='External sweep'

r=0
for i in list_rank1:
    Sheet1.cell(row=2+r,column=30).value=i
    r+=1
Sheet1['U2']=4778
for i in range(1,len(list_rank1)+1):
    if Sheet1.cell(row=1+i,column=30).value==1:
        Sheet1.cell(row=1+i,column=31).value='Internal outward interaction'
    if Sheet1.cell(row=1+i,column=30).value==-1:
        Sheet1.cell(row=1+i,column=31).value='External outward interaction'
    if Sheet1.cell(row=1+i,column=30).value==2:
        Sheet1.cell(row=1+i,column=31).value='External Ejection'
    if Sheet1.cell(row=1+i,column=30).value==-2:
        Sheet1.cell(row=1+i,column=31).value='Internal Ejection'
    if Sheet1.cell(row=1+i,column=30).value==3:
        Sheet1.cell(row=1+i,column=31).value=' External inward interaction'
    if Sheet1.cell(row=1+i,column=30).value==-3:
        Sheet1.cell(row=1+i,column=31).value='Internal inward interaction'
    if Sheet1.cell(row=1+i,column=30).value==4:
        Sheet1.cell(row=1+i,column=31).value='Internal sweep'
    if Sheet1.cell(row=1+i,column=30).value==-4:
        Sheet1.cell(row=1+i,column=31).value='External sweep'
    if Sheet1.cell(row=1+i,column=30).value==' ':
        Sheet1.cell(row=1+i,column=31).value=' '


Sheet1['P14']=list_rank1.count(1)
Sheet1['P15']=list_rank1.count(-1)
Sheet1['P16']=list_rank1.count(2)
Sheet1['P17']=list_rank1.count(-2)
Sheet1['P18']=list_rank1.count(3)
Sheet1['P19']=list_rank1.count(-3)
Sheet1['P20']=list_rank1.count(4)
Sheet1['P21']=list_rank1.count(-4)

for i in range(0,8):
    Sheet1.cell(row=3,column=22+i).value=None

wb.save("output_ranking.xlsx")





